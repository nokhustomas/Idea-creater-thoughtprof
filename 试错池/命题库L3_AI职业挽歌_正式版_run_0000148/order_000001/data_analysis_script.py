#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
data_analysis_script.py
=======================
From the cleaned collection (data_collection_report.xlsx, sheet 清洗后数据),
analyse and filter the top 5 occupations most impacted by AI.

A machine-check (assert-based validation) is embedded to guarantee the
analytical output is reproducible and consistent.

Output: data_analysis_report.xlsx with three sheets:
  - 受影响职业清单  (the final filtered list of 5 occupations)
  - 机器检查结果    (machine check results: each assertion PASS/FAIL)
  - 影响度评分      (per-occupation impact score breakdown)

The selection rule:
  impact_score = 0.5 * tasks_affected_pct - 0.2 * yoy_demand_change_pct
so that high task automation AND negative YoY demand both push the
occupation up the list.
"""
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

WORK = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(WORK, "data_collection_report.xlsx")
OUT = os.path.join(WORK, "data_analysis_report.xlsx")

TARGET_NUM = 5


def _style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="305496")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def load_cleaned():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["清洗后数据"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}
    data = []
    for r in rows[1:]:
        data.append({
            "occupation": r[idx["Occupation"]],
            "ai_tools_coverage": r[idx["AI Tools Coverage"]],
            "tasks_affected_pct": float(r[idx["Tasks Affected (%)"]]),
            "yoy_demand_change_pct": float(r[idx["YoY Demand Change (%)"]]),
            "data_source": r[idx["Data Source"]],
            "data_url": r[idx["Data URL"]],
        })
    return data


def analyse(data):
    # Simple weighted impact score
    for d in data:
        d["impact_score"] = round(0.5 * d["tasks_affected_pct"]
                                  - 0.2 * d["yoy_demand_change_pct"], 2)
    data.sort(key=lambda x: x["impact_score"], reverse=True)
    top = data[:TARGET_NUM]
    return data, top


def machine_check(top):
    """Embedded machine-check: a series of assertions on the output.

    Any failure raises AssertionError so the script exits non-zero and
    the failure is recorded in the 机器检查结果 sheet.
    """
    checks = []
    try:
        assert len(top) == TARGET_NUM, \
            f"必须正好 {TARGET_NUM} 个职业，实际 {len(top)}"
        checks.append(("count == 5", "PASS", f"{len(top)} occupations"))
    except AssertionError as e:
        checks.append(("count == 5", "FAIL", str(e)))

    try:
        names = [t["occupation"] for t in top]
        assert len(set(names)) == len(names), "职业名重复"
        checks.append(("unique occupation names", "PASS", f"{len(set(names))} unique"))
    except AssertionError as e:
        checks.append(("unique occupation names", "FAIL", str(e)))

    try:
        for t in top:
            assert 0 < t["tasks_affected_pct"] <= 100, \
                f"{t['occupation']} tasks_affected_pct 越界: {t['tasks_affected_pct']}"
        checks.append(("tasks_affected_pct in (0,100]", "PASS", "all valid"))
    except AssertionError as e:
        checks.append(("tasks_affected_pct in (0,100]", "FAIL", str(e)))

    try:
        for t in top:
            assert -100 <= t["yoy_demand_change_pct"] <= 100, \
                f"{t['occupation']} yoy_demand_change_pct 越界"
        checks.append(("yoy_demand_change_pct in [-100,100]", "PASS", "all valid"))
    except AssertionError as e:
        checks.append(("yoy_demand_change_pct in [-100,100]", "FAIL", str(e)))

    try:
        scores = [t["impact_score"] for t in top]
        assert scores == sorted(scores, reverse=True), "结果未按影响度降序"
        checks.append(("sorted by impact_score desc", "PASS", f"{scores}"))
    except AssertionError as e:
        checks.append(("sorted by impact_score desc", "FAIL", str(e)))

    try:
        full = [t["data_url"] for t in top]
        assert all(isinstance(u, str) and u.startswith("http") for u in full), \
            "数据 URL 缺失或非法"
        checks.append(("data_url present and http(s)", "PASS", "all 5 have URL"))
    except AssertionError as e:
        checks.append(("data_url present and http(s)", "FAIL", str(e)))

    try:
        for t in top:
            assert t["ai_tools_coverage"], \
                f"{t['occupation']} AI 工具覆盖字段为空"
        checks.append(("ai_tools_coverage not empty", "PASS", "all non-empty"))
    except AssertionError as e:
        checks.append(("ai_tools_coverage not empty", "FAIL", str(e)))

    try:
        tools_union = set()
        for t in top:
            for tok in t["ai_tools_coverage"].split():
                tools_union.add(tok)
        assert len(tools_union) >= 5, \
            f"AI 工具覆盖种类不足 5，实 {len(tools_union)}"
        checks.append((">=5 distinct AI tools mentioned", "PASS",
                       f"{len(tools_union)} tools"))
    except AssertionError as e:
        checks.append((">=5 distinct AI tools mentioned", "FAIL", str(e)))

    return checks


def build_report(top, checks, all_data):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "受影响职业清单"
    ws1.append(["Rank", "Occupation", "AI Tools Coverage",
                "Tasks Affected (%)", "YoY Demand Change (%)",
                "Impact Score", "Data Source", "Data URL"])
    for i, t in enumerate(top, 1):
        ws1.append([i, t["occupation"], t["ai_tools_coverage"],
                    t["tasks_affected_pct"], t["yoy_demand_change_pct"],
                    t["impact_score"], t["data_source"], t["data_url"]])
    _style_header(ws1)
    for col, w in zip("ABCDEFGH", [6, 36, 36, 18, 22, 14, 40, 70]):
        ws1.column_dimensions[col].width = w

    ws2 = wb.create_sheet("机器检查结果")
    ws2.append(["Check", "Status", "Detail"])
    for c in checks:
        ws2.append(list(c))
    _style_header(ws2)
    for col, w in zip("ABC", [40, 8, 60]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("影响度评分")
    ws3.append(["Occupation", "Tasks Affected (%)",
                "YoY Demand Change (%)", "Impact Score",
                "Formula", "Rank"])
    full_sorted = sorted(all_data, key=lambda x: x["impact_score"], reverse=True)
    rank_map = {d["occupation"]: i + 1 for i, d in enumerate(full_sorted)}
    for d in full_sorted:
        ws3.append([d["occupation"], d["tasks_affected_pct"],
                    d["yoy_demand_change_pct"], d["impact_score"],
                    "0.5*tasks - 0.2*(-yoy)",
                    rank_map[d["occupation"]]])
    _style_header(ws3)
    for col, w in zip("ABCDEF", [36, 18, 22, 14, 22, 6]):
        ws3.column_dimensions[col].width = w

    wb.save(OUT)
    print("[OK] wrote", OUT, " size=", os.path.getsize(OUT), "bytes")


def main():
    if not os.path.exists(SRC):
        print(f"[FAIL] missing {SRC}", file=sys.stderr)
        sys.exit(1)
    data = load_cleaned()
    all_data, top = analyse(data)
    checks = machine_check(top)
    build_report(top, checks, all_data)
    fails = [c for c in checks if c[1] == "FAIL"]
    if fails:
        print("[FAIL] machine check failed:", fails, file=sys.stderr)
        sys.exit(1)
    print(f"[OK] {len(top)} occupations, {len(checks)} checks PASS")


if __name__ == "__main__":
    main()